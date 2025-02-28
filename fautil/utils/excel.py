"""
Excel处理模块

提供Excel文件的导入和导出功能。
"""

import io
from typing import Any, Dict, List, Optional, Type, TypeVar, Union, cast

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

T = TypeVar("T", bound=BaseModel)


class ExcelExporter:
    """
    Excel导出工具类

    将数据导出为Excel文件。
    """

    def __init__(
        self,
        header_font: Optional[Dict[str, Any]] = None,
        header_alignment: Optional[Dict[str, Any]] = None,
        cell_alignment: Optional[Dict[str, Any]] = None,
    ):
        """
        初始化Excel导出器

        Args:
            header_font: 表头字体样式
            header_alignment: 表头对齐方式
            cell_alignment: 单元格对齐方式
        """
        self.header_font = header_font or {
            "name": "Arial",
            "size": 12,
            "bold": True,
        }
        self.header_alignment = header_alignment or {
            "horizontal": "center",
            "vertical": "center",
        }
        self.cell_alignment = cell_alignment or {
            "horizontal": "left",
            "vertical": "center",
        }

    def export_models(
        self,
        data: List[BaseModel],
        headers: Optional[Dict[str, str]] = None,
        sheet_name: str = "Sheet1",
    ) -> bytes:
        """
        导出Pydantic模型列表为Excel

        Args:
            data: 模型列表
            headers: 表头映射，键为模型字段名，值为表头显示名
            sheet_name: 工作表名称

        Returns:
            bytes: Excel文件的二进制数据
        """
        if not data:
            return self._create_empty_workbook(sheet_name)

        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        if ws:
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # 获取模型的所有字段
        sample = data[0]
        model_fields = sample.model_fields.keys()

        # 如果没有提供表头映射，使用字段名作为表头
        if not headers:
            headers = {field: field for field in model_fields}

        # 写入表头
        for col_idx, field in enumerate(headers.keys(), start=1):
            cell = ws.cell(row=1, column=col_idx, value=headers[field])
            cell.font = Font(**self.header_font)
            cell.alignment = Alignment(**self.header_alignment)

        # 写入数据
        for row_idx, item in enumerate(data, start=2):
            item_dict = item.model_dump()
            for col_idx, field in enumerate(headers.keys(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=item_dict.get(field))
                cell.alignment = Alignment(**self.cell_alignment)

        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        # 返回Excel文件的二进制数据
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_dicts(
        self,
        data: List[Dict[str, Any]],
        headers: Dict[str, str],
        sheet_name: str = "Sheet1",
    ) -> bytes:
        """
        导出字典列表为Excel

        Args:
            data: 字典列表
            headers: 表头映射，键为字典键，值为表头显示名
            sheet_name: 工作表名称

        Returns:
            bytes: Excel文件的二进制数据
        """
        if not data:
            return self._create_empty_workbook(sheet_name)

        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        if ws:
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # 写入表头
        for col_idx, field in enumerate(headers.keys(), start=1):
            cell = ws.cell(row=1, column=col_idx, value=headers[field])
            cell.font = Font(**self.header_font)
            cell.alignment = Alignment(**self.header_alignment)

        # 写入数据
        for row_idx, item in enumerate(data, start=2):
            for col_idx, field in enumerate(headers.keys(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=item.get(field))
                cell.alignment = Alignment(**self.cell_alignment)

        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        # 返回Excel文件的二进制数据
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _create_empty_workbook(self, sheet_name: str = "Sheet1") -> bytes:
        """
        创建空工作簿

        Args:
            sheet_name: 工作表名称

        Returns:
            bytes: Excel文件的二进制数据
        """
        wb = Workbook()
        ws = wb.active
        if ws:
            ws.title = sheet_name

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class ExcelImporter:
    """
    Excel导入工具类

    从Excel文件导入数据。
    """

    def __init__(self, skip_empty_rows: bool = True):
        """
        初始化Excel导入器

        Args:
            skip_empty_rows: 是否跳过空行
        """
        self.skip_empty_rows = skip_empty_rows

    def import_to_models(
        self,
        excel_data: Union[bytes, str],
        model_class: Type[T],
        field_mapping: Optional[Dict[str, str]] = None,
        sheet_name: Optional[str] = None,
        has_header: bool = True,
    ) -> List[T]:
        """
        从Excel导入数据到Pydantic模型

        Args:
            excel_data: Excel文件数据或文件路径
            model_class: Pydantic模型类
            field_mapping: 字段映射，键为Excel表头，值为模型字段名
            sheet_name: 工作表名称，如果为None则使用第一个工作表
            has_header: 是否有表头行

        Returns:
            List[T]: 模型实例列表
        """
        # 加载工作簿
        if isinstance(excel_data, bytes):
            wb = openpyxl.load_workbook(io.BytesIO(excel_data), data_only=True)
        else:
            wb = openpyxl.load_workbook(excel_data, data_only=True)

        # 获取工作表
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 读取数据
        return self._process_worksheet(ws, model_class, field_mapping, has_header)

    def import_to_dicts(
        self,
        excel_data: Union[bytes, str],
        field_mapping: Optional[Dict[str, str]] = None,
        sheet_name: Optional[str] = None,
        has_header: bool = True,
    ) -> List[Dict[str, Any]]:
        """
        从Excel导入数据到字典列表

        Args:
            excel_data: Excel文件数据或文件路径
            field_mapping: 字段映射，键为Excel表头，值为字典键
            sheet_name: 工作表名称，如果为None则使用第一个工作表
            has_header: 是否有表头行

        Returns:
            List[Dict[str, Any]]: 字典列表
        """
        # 加载工作簿
        if isinstance(excel_data, bytes):
            wb = openpyxl.load_workbook(io.BytesIO(excel_data), data_only=True)
        else:
            wb = openpyxl.load_workbook(excel_data, data_only=True)

        # 获取工作表
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 读取数据
        rows = list(ws.rows)
        if not rows:
            return []

        result = []
        headers = []

        # 处理表头
        if has_header:
            header_row = rows[0]
            headers = [cell.value for cell in header_row]
            data_rows = rows[1:]
        else:
            # 如果没有表头，使用列索引作为键
            headers = [f"column_{i}" for i in range(len(rows[0]))]
            data_rows = rows

        # 应用字段映射
        if field_mapping:
            # 创建反向映射
            reverse_mapping = {v: k for k, v in field_mapping.items()}
            # 将表头转换为模型字段名
            headers = [reverse_mapping.get(h, h) for h in headers]

        # 处理数据行
        for row in data_rows:
            row_data = {}
            is_empty = True

            for i, cell in enumerate(row):
                if i < len(headers):
                    value = cell.value
                    if value is not None:
                        is_empty = False
                    row_data[headers[i]] = value

            if not is_empty or not self.skip_empty_rows:
                result.append(row_data)

        return result

    def _process_worksheet(
        self,
        ws: Worksheet,
        model_class: Type[T],
        field_mapping: Optional[Dict[str, str]],
        has_header: bool,
    ) -> List[T]:
        """
        处理工作表数据

        Args:
            ws: 工作表
            model_class: Pydantic模型类
            field_mapping: 字段映射
            has_header: 是否有表头

        Returns:
            List[T]: 模型实例列表
        """
        # 先转换为字典列表
        dicts = self.import_to_dicts(
            excel_data=cast(bytes, ws),  # 类型转换只是为了类型检查，实际不会使用
            field_mapping=field_mapping,
            has_header=has_header,
        )

        # 转换为模型实例
        result = []
        for data in dicts:
            try:
                model = model_class(**data)
                result.append(model)
            except Exception as e:
                # 可以在这里处理验证错误
                raise ValueError(f"数据验证失败: {str(e)}")

        return result
